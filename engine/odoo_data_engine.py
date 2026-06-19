"""
Odoo ERP Data Migration Engine
================================
Universal AI-powered cleaning and column-mapping pipeline.
Converts arbitrary client spreadsheets into Odoo-ready import files.

Author  : Senior Python / AI Data Engineer
Version : 1.0.0
"""

from __future__ import annotations

import json
import logging
import re
import time
from pathlib import Path
from typing import Any

import anthropic
import os
import pandas as pd
import urllib.error
import urllib.request
from openpyxl.styles import PatternFill


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s – %(message)s",
)
logger = logging.getLogger("odoo_engine")


# ===========================================================================
# 1. ODOO TARGET SCHEMAS
#    Defines every field the templates require, with metadata used for
#    validation and default-filling.
# ===========================================================================

CUSTOMER_SCHEMA: dict[str, dict] = {
    "Title":                      {"required": False, "default": None, "dtype": "str"},
    "Name":                       {"required": True,  "default": None, "dtype": "str"},
    "Job Position":               {"required": False, "default": None, "dtype": "str"},
    "Mobile":                     {"required": False, "default": None, "dtype": "str"},
    "Phone":                      {"required": False, "default": None, "dtype": "str"},
    "Email":                      {"required": False, "default": None, "dtype": "str"},
    "Street":                     {"required": False, "default": None, "dtype": "str"},
    "Street2":                    {"required": False, "default": None, "dtype": "str"},
    "City":                       {"required": False, "default": None, "dtype": "str"},
    "State":                      {"required": False, "default": None, "dtype": "str"},
    "Zip":                        {"required": False, "default": None, "dtype": "str"},
    "Country":                    {"required": False, "default": None, "dtype": "str"},
    "Website":                    {"required": False, "default": None, "dtype": "str"},
    "Company":                    {"required": False, "default": None, "dtype": "str"},
    "Is a Company":               {"required": False, "default": True,  "dtype": "bool"},
    "Reference":                  {"required": False, "default": None, "dtype": "str"},
    "Credit Limit":               {"required": False, "default": 0,    "dtype": "float"},
    "Branch":                     {"required": False, "default": None, "dtype": "str"},
    "Customer Rank":              {"required": False, "default": 1,    "dtype": "int"},
    "Salesperson":                {"required": False, "default": None, "dtype": "str"},
    "Category":                   {"required": False, "default": None, "dtype": "str"},
    "Sub Category":               {"required": False, "default": None, "dtype": "str"},
    "Note":                       {"required": False, "default": None, "dtype": "str"},
    "Account Receivable":         {"required": False, "default": None, "dtype": "str"},
    "Account Payable":            {"required": False, "default": None, "dtype": "str"},
    "contacts / name":            {"required": False, "default": None, "dtype": "str"},
    "contacts / email":           {"required": False, "default": None, "dtype": "str"},
    "contacts / title":           {"required": False, "default": None, "dtype": "str"},
    "contacts / phone":           {"required": False, "default": None, "dtype": "str"},
    "contacts / mobile":          {"required": False, "default": None, "dtype": "str"},
    "contacts / Job Title":       {"required": False, "default": None, "dtype": "str"},
}

VENDOR_SCHEMA: dict[str, dict] = {
    "Vendor Name":                {"required": True,  "default": None, "dtype": "str"},
    "Street":                     {"required": False, "default": None, "dtype": "str"},
    "Street2":                    {"required": False, "default": None, "dtype": "str"},
    "City":                       {"required": False, "default": None, "dtype": "str"},
    "State":                      {"required": False, "default": None, "dtype": "str"},
    "Zip":                        {"required": False, "default": None, "dtype": "str"},
    "Country":                    {"required": False, "default": None, "dtype": "str"},
    "Tax ID":                     {"required": False, "default": None, "dtype": "str"},
    "Phone":                      {"required": False, "default": None, "dtype": "str"},
    "Mobile":                     {"required": False, "default": None, "dtype": "str"},
    "Email":                      {"required": False, "default": None, "dtype": "str"},
    "Website":                    {"required": False, "default": None, "dtype": "str"},
    "Is a Company":               {"required": False, "default": True,  "dtype": "bool"},
    "Reference":                  {"required": False, "default": None, "dtype": "str"},
    "Supplier Rank":              {"required": False, "default": 1,    "dtype": "int"},
    "Address Type":               {"required": False, "default": None, "dtype": "str"},
    "Tags":                       {"required": False, "default": None, "dtype": "str"},
    "Contact/Name":               {"required": False, "default": None, "dtype": "str"},
    "Contact/Title":              {"required": False, "default": None, "dtype": "str"},
    "Contact/Job Position":       {"required": False, "default": None, "dtype": "str"},
    "Contact/Email":              {"required": False, "default": None, "dtype": "str"},
    "Contact/Phone":              {"required": False, "default": None, "dtype": "str"},
    "Contact/Mobile":             {"required": False, "default": None, "dtype": "str"},
}

SCHEMA_MAP = {"customer": CUSTOMER_SCHEMA, "vendor": VENDOR_SCHEMA}
MANDATORY_FIELD = {"customer": "Name", "vendor": "Vendor Name"}


# ===========================================================================
# 1c. UNIVERSAL FILE STRUCTURE NORMALISER
#     Detects non-flat layouts and reshapes them into a standard flat table.
# ===========================================================================

def ai_normalise_structure(
    df: pd.DataFrame,
    source_path: Path,
    client: anthropic.Anthropic,
) -> pd.DataFrame:
    """
    Detects whether a loaded DataFrame has a non-standard structure
    (block records, vertical layout, multi-row headers, mixed metadata rows, etc.)
    and uses the AI to produce a reshape plan, then executes it in pure Pandas.

    For standard flat tables (one header row, one row per record) this returns
    the DataFrame unchanged — zero API cost.

    Handles:
    - Block structures  : multiple raw rows per record (e.g. CSS vendor file)
    - Vertical layouts  : fields as rows, records as columns
    - Multi-row headers : 2-3 header rows that need collapsing
    - Junk preamble     : title/logo rows before the real data starts
    - Mixed rows        : metadata/summary rows interspersed with data rows
    """
    # ── Quick pre-check: is this already a standard flat table? ──
    # Heuristic: if row 0 looks like a header (mostly string tokens, no repeating
    # label patterns like "Name:", "A/C Ref:") → flat table, return as-is.
    if _is_flat_table(df):
        return df

    logger.info("Non-flat structure detected in '%s' — running AI structure analyser", source_path.name)

    # Send a compact snapshot of the raw file to the AI:
    # first 12 rows × all columns, serialised as a JSON array of arrays.
    # 12 rows is enough to see the repeating pattern without blowing the token budget.
    snapshot_rows = min(12, len(df))
    snapshot = df.head(snapshot_rows).fillna("").astype(str).values.tolist()

    system_prompt = (
        "You are an expert data engineer specialising in spreadsheet normalisation. "
        "Return ONLY valid JSON — no prose, no markdown, no code fences."
    )

    user_prompt = f"""A spreadsheet has been loaded into a 2D array. The structure is NOT a standard
flat table (one header row followed by one row per record). You must analyse the pattern
and return a JSON reshape plan.

RAW DATA SNAPSHOT (first {snapshot_rows} rows, all columns):
{json.dumps(snapshot, indent=2)}

Analyse the repeating pattern carefully. Common non-flat structures:
1. BLOCK  — multiple raw rows form one logical record.
   e.g. Row A = "A/C Ref: | REF001 | Name: | Vendor Name"
        Row B = "0803... | Street | Sor | email@... | ..."
   Each pair of rows = one vendor record.

2. VERTICAL — fields are rows, records are columns.
   e.g. Row 0 = "Name | Alice | Bob | Carol"
        Row 1 = "Phone | 080... | 070... | 090..."

3. MULTI_HEADER — 2-3 rows together form the column header.
   e.g. Row 0 = "Contact | | Address | |"
        Row 1 = "Name | Phone | Street | City | State"

4. PREAMBLE — junk rows (title, logo text, blank) before the real data.
   e.g. Row 0 = "Agary Publishing — Customer List 2024"
        Row 1 = "" (blank)
        Row 2 = "Name | Phone | Address | ..."  ← real header

Return a JSON object with this structure:
{{
  "structure_type": "FLAT" | "BLOCK" | "VERTICAL" | "MULTI_HEADER" | "PREAMBLE" | "MIXED",
  "description": "one sentence describing what you see",
  "reshape_plan": {{

    For FLAT (already standard — return as-is):
      {{"action": "none"}}

    For BLOCK:
      {{
        "action": "block_to_rows",
        "rows_per_record": <int>,        // how many raw rows form one record
        "header_row_index": <int|null>,  // index of a header row if present, else null
        "skip_rows": [<int>, ...],       // raw row indices to skip (e.g. file header)
        "record_start_signal": {{        // how to identify the FIRST row of each block
          "col": <int>,                  // column index
          "value": "<string>"            // value that marks the start of a new record
        }},
        "field_map": [                   // how to read each field from the block rows
          {{
            "field": "<human name>",     // what this field represents
            "row_offset": <int>,         // 0 = first row of block, 1 = second row, etc.
            "col": <int>                 // column index within that row
          }},
          ...
        ]
      }}

    For VERTICAL:
      {{
        "action": "transpose",
        "field_name_col": <int>,   // column that contains field names
        "data_start_col": <int>    // first column that contains record data
      }}

    For MULTI_HEADER:
      {{
        "action": "collapse_headers",
        "header_rows": [<int>, ...],   // row indices that together form the header
        "data_start_row": <int>        // first row of actual data
      }}

    For PREAMBLE:
      {{
        "action": "skip_preamble",
        "header_row": <int>,    // row index of the real header
        "data_start_row": <int> // first row of actual data
      }}

    For MIXED:
      {{
        "action": "filter_rows",
        "header_row": <int>,
        "data_row_signal": {{   // how to identify a real data row vs junk
          "col": <int>,
          "pattern": "<regex or keyword that data rows match>"
        }}
      }}
  }}
}}

Return ONLY the JSON object.
"""

    max_retries = 3
    for attempt in range(max_retries):
        try:
            raw_text = _complete_with_fallback(client, system_prompt, user_prompt, max_tokens=1500)
            raw_text = re.sub(r"^```[a-z]*\n?|```$", "", raw_text, flags=re.MULTILINE).strip()
            plan = json.loads(raw_text)
            logger.info("Structure analysis: type=%s | %s",
                        plan.get("structure_type"), plan.get("description"))
            return _execute_reshape_plan(df, plan, source_path)
        except (json.JSONDecodeError, _LLMUnavailable) as exc:
            logger.warning("Structure analysis attempt %d failed: %s", attempt + 1, exc)
            time.sleep(2 ** attempt)

    logger.warning("Structure analysis failed — returning raw DataFrame as-is")
    return df


def _is_flat_table(df: pd.DataFrame) -> bool:
    """
    Quick heuristic: does this DataFrame look like a standard flat table?
    Returns True if it does (skip AI call), False if it needs AI analysis.

    Signals of a NON-flat table:
    - A label like "A/C Ref:", "Name:", "Tel:", "Code:" repeats many times in col 0
      (block structure marker)
    - More than 30% of rows in col 0 are blank (block structure or vertical)
    - Col 0 row 0 is blank (preamble or vertical layout)
    """
    if df.empty or len(df) < 3:
        return True

    col0 = df.iloc[:, 0].fillna("").astype(str).str.strip()

    # Signal 1: repeating label pattern in col 0 (block structure)
    label_pattern = re.compile(r"^(a/c ref:?|name:|tel:?|code:?|ref:?|id:?|telephone|a/c ref)$", re.IGNORECASE)
    label_count = sum(1 for v in col0 if label_pattern.match(v))
    if label_count >= 3:
        return False

    # Signal 2: >30% blank in col 0 (block or vertical)
    blank_count = sum(1 for v in col0 if not v)
    if blank_count / len(col0) > 0.30:
        return False

    # Signal 3: first row looks like field labels not column headers
    # (very short values, all different, look like "Name", "Phone" etc.)
    row0 = df.iloc[0].fillna("").astype(str).str.strip()
    if row0.iloc[0] == "" and df.shape[1] > 3:
        return False

    return True


def _execute_reshape_plan(
    df: pd.DataFrame,
    plan: dict,
    source_path: Path,
) -> pd.DataFrame:
    """
    Execute the reshape plan returned by the AI.
    Pure Pandas — no additional API calls.
    """
    action = plan.get("reshape_plan", {}).get("action", "none")

    if action == "none":
        return df

    rp = plan["reshape_plan"]

    # ── BLOCK: multiple rows per record ──
    if action == "block_to_rows":
        return _reshape_block(df, rp)

    # ── VERTICAL: transpose rows↔cols ──
    if action == "transpose":
        field_col  = rp.get("field_name_col", 0)
        data_start = rp.get("data_start_col", 1)
        transposed = df.iloc[:, data_start:].T.copy()
        transposed.columns = df.iloc[:, field_col].fillna("").astype(str).tolist()
        transposed.reset_index(drop=True, inplace=True)
        transposed.dropna(how="all", inplace=True)
        logger.info("Transposed vertical layout → %d rows × %d cols", len(transposed), len(transposed.columns))
        return transposed

    # ── MULTI_HEADER: collapse header rows ──
    if action == "collapse_headers":
        header_rows = rp.get("header_rows", [0])
        data_start  = rp.get("data_start_row", len(header_rows))
        # Combine header rows by joining non-empty values
        combined_header = []
        for col_idx in range(df.shape[1]):
            parts = [
                str(df.iloc[r, col_idx]).strip()
                for r in header_rows
                if str(df.iloc[r, col_idx]).strip() not in ("", "nan")
            ]
            combined_header.append(" / ".join(parts) if parts else f"col_{col_idx}")
        result = df.iloc[data_start:].copy()
        result.columns = combined_header[:len(result.columns)]
        result.reset_index(drop=True, inplace=True)
        logger.info("Collapsed %d header rows → %d data rows", len(header_rows), len(result))
        return result

    # ── PREAMBLE: skip junk rows before real data ──
    if action == "skip_preamble":
        header_row = rp.get("header_row", 0)
        data_start = rp.get("data_start_row", header_row + 1)
        result = df.iloc[data_start:].copy()
        result.columns = df.iloc[header_row].fillna("").astype(str).str.strip().tolist()
        result.reset_index(drop=True, inplace=True)
        result.dropna(how="all", inplace=True)
        logger.info("Skipped preamble — %d data rows", len(result))
        return result

    # ── MIXED: filter out non-data rows ──
    if action == "filter_rows":
        header_row  = rp.get("header_row", 0)
        signal      = rp.get("data_row_signal", {})
        sig_col     = signal.get("col", 0)
        sig_pattern = signal.get("pattern", "")
        result = df.iloc[header_row + 1:].copy()
        result.columns = df.iloc[header_row].fillna("").astype(str).str.strip().tolist()
        if sig_pattern:
            mask = result.iloc[:, sig_col].astype(str).str.contains(sig_pattern, na=False, regex=True)
            result = result[mask]
        result.reset_index(drop=True, inplace=True)
        logger.info("Filtered mixed rows → %d data rows", len(result))
        return result

    logger.warning("Unknown reshape action '%s' — returning raw DataFrame", action)
    return df


def _reshape_block(df: pd.DataFrame, rp: dict) -> pd.DataFrame:
    """
    Reshape a block-structured DataFrame where N raw rows = 1 record.
    Uses the AI-provided field_map to extract each field from the correct
    row offset and column within each block.
    """
    signal_col = rp.get("record_start_signal", {}).get("col", 0)
    signal_val = rp.get("record_start_signal", {}).get("value", "")
    skip_rows  = set(rp.get("skip_rows", []))
    field_map  = rp.get("field_map", [])

    if not field_map:
        logger.warning("Block reshape: no field_map provided — returning raw")
        return df

    # Find all block start indices
    block_starts = []
    for i in range(len(df)):
        if i in skip_rows:
            continue
        cell = str(df.iloc[i, signal_col]).strip()
        if cell == signal_val:
            block_starts.append(i)

    if not block_starts:
        logger.warning("Block reshape: no block start rows found for signal %r=%r", signal_col, signal_val)
        return df

    # Extract each block into a record
    records = []
    col_names = [f["field"] for f in field_map]

    for start in block_starts:
        record = {}
        for fm in field_map:
            row_offset = fm.get("row_offset", 0)
            col_idx    = fm.get("col", 0)
            raw_row    = start + row_offset
            if raw_row < len(df) and col_idx < df.shape[1]:
                val = str(df.iloc[raw_row, col_idx]).strip()
                record[fm["field"]] = val if val not in ("", "nan") else None
            else:
                record[fm["field"]] = None
        records.append(record)

    result = pd.DataFrame(records, columns=col_names)
    result.dropna(how="all", inplace=True)
    result.reset_index(drop=True, inplace=True)
    logger.info("Block reshape: %d blocks → %d records × %d fields",
                len(block_starts), len(result), len(col_names))
    return result


# ===========================================================================
# 2b. LLM PROVIDER LAYER  –  Claude first, free OpenRouter model as fallback
# ===========================================================================
# The Anthropic free tier rate-limits hard (requests/minute AND tokens/day).
# Rather than let a 429 stall or kill the run, every AI call in this module
# goes through _complete_with_fallback(): the moment Anthropic returns a
# rate-limit response, we re-send the SAME prompt to a free model on
# OpenRouter (https://openrouter.ai – OpenAI-compatible REST API, no cost
# for "...:free" models) instead of spinning and burning more Anthropic quota.
#
# Configure via environment variables:
#   OPENROUTER_API_KEY  – your OpenRouter key (free signup at openrouter.ai/keys)
#   OPENROUTER_MODEL    – defaults to a free Llama 3.3 70B slug; override with
#                         any "...:free" model id from openrouter.ai/models
#
# Without OPENROUTER_API_KEY set, the fallback is a no-op and behaviour is
# identical to before (Anthropic-only, with backoff/retry).

OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
OPENROUTER_MODEL   = os.getenv("OPENROUTER_MODEL", "meta-llama/llama-3.3-70b-instruct:free")
OPENROUTER_URL     = "https://openrouter.ai/api/v1/chat/completions"

# Gemini sits between Claude and OpenRouter: better quality than OpenRouter's
# free models, and (on the free Gemini tier) gemini-3.1-flash-lite gives a
# much higher daily quota (500 RPD) than e.g. gemini-2.5-flash (20 RPD) — see
# https://ai.dev/rate-limit for current per-model limits on your key.
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
GEMINI_MODEL   = os.getenv("GEMINI_MODEL", "gemini-3.1-flash-lite")
GEMINI_URL     = "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"

# How many times the Anthropic SDK retries a 429 internally before giving up
# and letting _complete_with_fallback() switch to the fallback chain. Claude's
# output quality is the best of the three, so it's worth a few quick SDK-level
# retries first rather than dropping down on the very first rate-limit
# response — but capped, so a sustained rate limit still fails over instead
# of retrying forever.
ANTHROPIC_MAX_RETRIES = int(os.getenv("ANTHROPIC_MAX_RETRIES", "3"))

# Per-provider request pacing. RATE_LIMIT_RPM previously existed in .env but
# was never actually read anywhere — every call fired as fast as the loop
# could go, relying entirely on Anthropic's SDK retrying its own 429s. Each
# provider in the fallback chain has a different real limit (Claude's free
# tier is ~5 RPM; gemini-3.1-flash-lite's free tier is 15 RPM), so pacing them
# all to one shared rate would throttle a faster fallback down to the
# slowest provider's pace. Each gets its own independent minimum interval.
ANTHROPIC_RATE_LIMIT_RPM = float(os.getenv("RATE_LIMIT_RPM") or 5)
GEMINI_RATE_LIMIT_RPM    = float(os.getenv("GEMINI_RATE_LIMIT_RPM") or 15)

_last_call_at: dict[str, float] = {}


def _pace_call(provider: str, rpm: float) -> None:
    """Block until at least 60/rpm seconds have passed since the last call to this provider."""
    if rpm <= 0:
        return
    min_interval = 60.0 / rpm
    elapsed = time.monotonic() - _last_call_at.get(provider, 0.0)
    if elapsed < min_interval:
        time.sleep(min_interval - elapsed)
    _last_call_at[provider] = time.monotonic()


class _LLMUnavailable(Exception):
    """Raised when neither Anthropic nor the OpenRouter fallback could complete a request."""


def _is_rate_limited(exc: anthropic.APIError) -> bool:
    """True for HTTP 429 (rate-limit / quota-exceeded) responses — transient, worth retrying later."""
    if isinstance(exc, anthropic.RateLimitError):
        return True
    return getattr(exc, "status_code", None) == 429


def _is_credits_exhausted(exc: anthropic.APIError) -> bool:
    """
    True for "credit balance too low" — comes back as a 400
    invalid_request_error, not a 429, so status code alone won't catch it.
    Unlike a rate limit, this won't clear up on its own (needs a human to add
    billing), so the caller uses this to stop attempting Claude entirely for
    the rest of the run instead of repeatedly paying the pacing wait only to
    fail the same way every time.
    """
    return "credit balance" in str(exc).lower()


def _is_provider_unavailable_error(exc: anthropic.APIError) -> bool:
    """
    True for errors that mean Claude itself can't serve this request right
    now (rate-limited, or out of credits) rather than the request being
    malformed — these are worth falling back to Gemini/OpenRouter for,
    since the same prompt would likely succeed elsewhere.
    """
    return _is_rate_limited(exc) or _is_credits_exhausted(exc)


# Set once a "credit balance too low" response is seen — skips Claude
# entirely for the rest of this process's life instead of re-attempting (and
# re-paying the pacing wait for) a call guaranteed to fail the same way again.
_anthropic_credits_exhausted = False


def _retry_after_seconds(exc: anthropic.APIError) -> float | None:
    """Read the Retry-After header off a rate-limit error, if the SDK exposes it."""
    response = getattr(exc, "response", None)
    headers  = getattr(response, "headers", None)
    value    = headers.get("retry-after") if headers else None
    if value:
        try:
            return float(value)
        except ValueError:
            pass
    return None


def _call_gemini(system_prompt: str, user_prompt: str, max_tokens: int) -> str | None:
    """
    Run the same prompt against Gemini.
    Returns the raw response text, or None if Gemini is not configured or the
    request fails for any reason (including the free tier being exhausted).
    """
    if not GEMINI_API_KEY:
        return None

    payload = json.dumps({
        "systemInstruction": {"parts": [{"text": system_prompt}]},
        "contents": [{"parts": [{"text": user_prompt}]}],
        "generationConfig": {
            "maxOutputTokens": max_tokens,
            # Without this, Gemini 2.5+ "thinking" models can spend the whole
            # max_tokens budget on internal reasoning and return no visible
            # text at all (finishReason MAX_TOKENS, empty content parts).
            "thinkingConfig": {"thinkingBudget": 0},
        },
    }).encode("utf-8")

    url = GEMINI_URL.format(model=GEMINI_MODEL, key=GEMINI_API_KEY)
    req = urllib.request.Request(
        url,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            body = json.loads(resp.read().decode("utf-8"))
        return body["candidates"][0]["content"]["parts"][0]["text"].strip()
    except (urllib.error.URLError, KeyError, IndexError, json.JSONDecodeError) as exc:
        logger.warning("Gemini fallback request failed: %s", exc)
        return None


def _call_openrouter(system_prompt: str, user_prompt: str, max_tokens: int) -> str | None:
    """
    Run the same prompt against a free model on OpenRouter.
    Returns the raw response text, or None if OpenRouter is not configured
    or the request fails for any reason.
    """
    if not OPENROUTER_API_KEY:
        return None

    payload = json.dumps({
        "model": OPENROUTER_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_prompt},
        ],
        "max_tokens": max_tokens,
    }).encode("utf-8")

    req = urllib.request.Request(
        OPENROUTER_URL,
        data=payload,
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type":  "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            body = json.loads(resp.read().decode("utf-8"))
        return body["choices"][0]["message"]["content"].strip()
    except (urllib.error.URLError, KeyError, IndexError, json.JSONDecodeError) as exc:
        logger.warning("OpenRouter fallback request failed: %s", exc)
        return None


def _complete_with_fallback(
    client: anthropic.Anthropic,
    system_prompt: str,
    user_prompt: str,
    max_tokens: int,
) -> str:
    """
    Send one completion request to Claude.

    When Claude itself is unavailable (rate-limited, or the account is out
    of credits) falls through the chain Claude → Gemini → OpenRouter,
    stopping at the first one that succeeds, instead of failing the request
    outright. Once a "credit balance too low" response is seen, Claude is
    skipped entirely for the rest of this run — that failure won't clear up
    on its own, so there's no point re-paying the pacing wait to attempt (and
    re-fail) it on every subsequent call.

    Any other Anthropic error (a malformed request, say — one that would
    fail identically on every provider) or an unconfigured/failed fallback
    chain raises _LLMUnavailable so the caller's existing retry loop handles it.
    """
    global _anthropic_credits_exhausted
    exc: anthropic.APIError | None = None

    if not _anthropic_credits_exhausted:
        _pace_call("anthropic", ANTHROPIC_RATE_LIMIT_RPM)
        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=max_tokens,
                system=system_prompt,
                messages=[{"role": "user", "content": user_prompt}],
            )
            return response.content[0].text.strip()
        except anthropic.APIError as caught:
            if _is_credits_exhausted(caught):
                logger.error(
                    "Anthropic account is out of credits – skipping Claude for "
                    "the rest of this run, falling back to Gemini/OpenRouter only"
                )
                _anthropic_credits_exhausted = True
            elif not _is_rate_limited(caught):
                raise _LLMUnavailable(str(caught)) from caught
            exc = caught

    if exc is not None:
        logger.warning("Anthropic unavailable (%s) – trying Gemini fallback (%s) …", exc, GEMINI_MODEL)
    _pace_call("gemini", GEMINI_RATE_LIMIT_RPM)
    fallback_text = _call_gemini(system_prompt, user_prompt, max_tokens)
    if fallback_text is not None:
        logger.info("Gemini fallback succeeded – continuing pipeline")
        return fallback_text

    logger.warning(
        "Gemini fallback unavailable – trying OpenRouter fallback (%s) …",
        OPENROUTER_MODEL,
    )
    fallback_text = _call_openrouter(system_prompt, user_prompt, max_tokens)
    if fallback_text is not None:
        logger.info("OpenRouter fallback succeeded – continuing pipeline")
        return fallback_text

    if exc is not None:
        wait = _retry_after_seconds(exc)
        if wait:
            logger.warning(
                "Anthropic unavailable, no fallback available – sleeping %.0fs …", wait
            )
            time.sleep(wait)
        raise _LLMUnavailable(
            "Anthropic unavailable and no fallback (Gemini/OpenRouter) available"
        ) from exc
    raise _LLMUnavailable(
        "Anthropic credits exhausted and no fallback (Gemini/OpenRouter) available"
    )


# ===========================================================================
# 2. FILE LOADER  –  handles .xlsx / .xls / .csv with auto-header detection
# ===========================================================================

def load_raw_file(filepath: str | Path) -> pd.DataFrame:
    """
    Load a spreadsheet or CSV into a DataFrame.
    Handles merged header rows (like CSS vendor data) by scanning the
    first 5 rows to find the actual header row.
    """
    path = Path(filepath)
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls", ".xlsm"):
        # First pass: sniff which row looks like a header
        probe = pd.read_excel(path, header=None, nrows=10)
        header_row = _detect_header_row(probe)
        df = pd.read_excel(path, header=header_row, dtype=str)
    elif suffix == ".csv":
        probe = pd.read_csv(path, header=None, nrows=10)
        header_row = _detect_header_row(probe)
        df = pd.read_csv(path, header=header_row, dtype=str)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    # Drop fully-empty rows and columns
    df.dropna(how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)

    # Strip whitespace from column names and cell values
    df.columns = [str(c).strip() for c in df.columns]
    df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))

    # Drop Pandas auto-generated "Unnamed" columns
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

    logger.info("Loaded %d rows × %d cols from '%s' (header row %d)",
                len(df), len(df.columns), path.name, header_row)
    return df



def _detect_header_row(probe: pd.DataFrame) -> int:
    """
    Find the row index most likely to be a column header by counting
    non-null, non-numeric string cells per row.
    """
    best_row, best_score = 0, 0
    for i, row in probe.iterrows():
        score = sum(
            1 for v in row
            if isinstance(v, str) and v.strip()
            and not re.fullmatch(r"[\d\s.,-/]+", v.strip())
        )
        if score > best_score:
            best_score, best_row = score, int(i)
    return best_row


# ===========================================================================
# 3. AI COLUMN MAPPER  –  one LLM call, structured JSON output
# ===========================================================================

def ai_map_columns(
    raw_columns: list[str],
    target_schema: dict[str, dict],
    data_type: str,
    client: anthropic.Anthropic,
    sample_rows: list[dict] | None = None,
) -> tuple[dict[str, str | None], dict[str, list[str]]]:
    """
    Ask Claude to map raw column names → Odoo target column names, AND identify
    columns that need additional AI processing (address splitting, field cleaning).

    Returns:
        (mapping, flags)
        - mapping: { raw_col_name: odoo_field_name_or_None }
        - flags:   {
              "needs_address_split": [raw col names whose values combine street+city+state],
              "needs_field_clean":   [odoo field names whose sample values look cross-contaminated]
          }
    """
    target_fields = list(target_schema.keys())
    sample_str = ""
    if sample_rows:
        sample_str = "\n\nSample data rows (first 3):\n" + json.dumps(
            sample_rows[:3], indent=2, default=str
        )

    system_prompt = (
        "You are an expert Odoo ERP data migration specialist. "
        "You MUST return ONLY valid JSON – no prose, no code fences, no explanations. "
        "Your response must be a single JSON object."
    )

    user_prompt = f"""Map the raw spreadsheet columns below to Odoo {data_type} import fields,
and flag any columns that need additional AI processing.

RAW COLUMNS (from client spreadsheet):
{json.dumps(raw_columns, indent=2)}
{sample_str}

TARGET ODOO FIELDS (exact spelling required):
{json.dumps(target_fields, indent=2)}

MAPPING RULES:
1. Each KEY is a raw column name; VALUE is the best-matching Odoo field name (exact spelling
   from the target list), or null if no match.
2. Never invent field names – only use values from the target list or null.
3. One raw column → at most one Odoo field. Multiple raws must NOT share the same Odoo target.
4. Semantic matching required: "Client Name"→"Name", "Organisation"→"Vendor Name",
   "Tel"→"Phone", "Ref"→"Reference", "Area"→"City", "Supplier Rank"→"Supplier Rank".
5. For child/contact columns match appropriately (e.g. "child_ids/name"→"contacts / name").
6. If a raw column contains combined address data (street + city + state in one cell),
   map it to "Street" AND add its raw column name to "needs_address_split".

FLAG RULES:
- "needs_address_split": list raw column names where sample values combine multiple address
  parts in one cell (e.g. "12 Broad St, Lagos, Lagos State"). Only flag when clearly combined.
- "needs_field_clean": list the ODOO FIELD NAMES (post-mapping) where sample values look
  cross-contaminated. Examples: Phone column with "John Smith 0803..." (name before number);
  State column with city names instead of state names; Country with full addresses.
  Only flag fields where contamination is evident — do not flag clean columns.

Return a JSON object with EXACTLY this structure (no other keys):
{{
  "mapping": {{
    "RawColumnName": "OdooFieldName",
    "AnotherRawCol": null
  }},
  "flags": {{
    "needs_address_split": ["RawColWithCombinedAddress"],
    "needs_field_clean": ["Phone", "State"]
  }}
}}
"""

    max_retries = 3
    for attempt in range(max_retries):
        try:
            raw_text = _complete_with_fallback(client, system_prompt, user_prompt, max_tokens=1000)
            raw_text = re.sub(r"^```[a-z]*\n?|```$", "", raw_text, flags=re.MULTILINE).strip()
            result   = json.loads(raw_text)
            mapping  = result.get("mapping", {})
            flags    = result.get("flags", {
                "needs_address_split": [],
                "needs_field_clean":   [],
            })
            logger.info(
                "AI column mapping: %d mappings | address_split=%s | field_clean=%s",
                len(mapping),
                flags.get("needs_address_split", []),
                flags.get("needs_field_clean",   []),
            )
            return mapping, flags
        except (json.JSONDecodeError, _LLMUnavailable) as exc:
            logger.warning("AI mapping attempt %d failed: %s", attempt + 1, exc)
            time.sleep(2 ** attempt)

    logger.error("AI column mapping failed after %d attempts – returning empty map", max_retries)
    return {}, {"needs_address_split": [], "needs_field_clean": []}


# ===========================================================================
# 4. RULE-BASED PANDAS CLEANER  –  fast, deterministic, zero API cost
# ===========================================================================

_PHONE_RE = re.compile(r"[^\d+]")          # strip non-digit / non-plus
_EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")


def rule_based_clean(df: pd.DataFrame, data_type: str) -> pd.DataFrame:
    """
    Apply deterministic Pandas transformations that never need LLM calls:
    - Phone → numeric string, strip dashes/brackets/spaces
    - Email → lowercase, validate format
    - Name / Company → Title Case
    - Boolean fields → True/False
    - Country → extracted from any address column in the row; left blank only if no clue found
    - Reference → strip internal Odoo export prefixes
    - Credit Limit / Rank → numeric with sensible defaults
    """
    schema = SCHEMA_MAP[data_type]

    # --- Name fields: Title Case ---
    name_cols = [c for c in df.columns if c.lower() in ("name", "vendor name", "company")]
    for col in name_cols:
        df[col] = df[col].apply(_title_case)

    # --- Phone / Mobile: clean all phone-like columns independently ---
    # Both Phone and Mobile are preserved as separate Odoo fields.
    # Raw "phone" -> "Phone" and raw "mobile" -> "Mobile" are kept distinct.
    # If a field contains a person name instead of digits, _clean_phone returns None.
    phone_cols = [c for c in df.columns if any(k in c.lower() for k in ("phone", "mobile", "tel"))]
    for col in phone_cols:
        df[col] = df[col].apply(_clean_phone)

    # --- child_ids: rescue contact names hidden in boolean column ---
    # In Odoo exports, child_ids is exported as False when no child contact exists.
    # But sometimes the BA enters the contact person name directly in that cell
    # (e.g. "TIMOTHY EKENE IKELUGO"). We detect this and move it to child_ids/name.
    if "child_ids" in df.columns:
        child_name_col = next(
            (c for c in df.columns if c.lower() in ("child_ids/name", "contacts / name")),
            None
        )
        def _rescue_child_id(row: pd.Series) -> pd.Series:
            val = row.get("child_ids", "")
            if not isinstance(val, str):
                return row
            v = val.strip()
            # "False" / empty / purely numeric = real Odoo boolean export, ignore
            if v.lower() in ("false", "true", "", "0", "1") or re.fullmatch(r"\d+", v):
                return row
            # Looks like an actual name: move it to child_ids/name if that cell is empty
            if child_name_col and (pd.isna(row.get(child_name_col)) or str(row.get(child_name_col, "")).strip() == ""):
                row = row.copy()
                row[child_name_col] = v.title()
            row["child_ids"] = None   # clear the mis-used cell
            return row

        if child_name_col:
            df = df.apply(_rescue_child_id, axis=1)
            logger.info("Rescued child_ids contact names into '%s'", child_name_col)
        # Drop the raw child_ids column — it is not an Odoo import field
        df.drop(columns=["child_ids"], inplace=True, errors="ignore")

    # --- Email: lowercase + validate ---
    email_cols = [c for c in df.columns if "email" in c.lower() or "e-mail" in c.lower()]
    for col in email_cols:
        df[col] = df[col].apply(_clean_email)

    # --- Website: lowercase ---
    web_cols = [c for c in df.columns if "website" in c.lower() or "web" in c.lower()]
    for col in web_cols:
        df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)

    # --- Is a Company: coerce to boolean if column exists ---
    company_cols = [c for c in df.columns if "is a company" in c.lower()]
    for col in company_cols:
        df[col] = df[col].apply(_coerce_bool)

    # --- Is a Company: INFER from name when column is absent ---
    # This handles files like FZR raw data that have no Is a Company column.
    # We look for the name column and derive the boolean from the name value.
    if not company_cols:
        name_col = next(
            (c for c in df.columns if c.lower() in ("name", "vendor name", "description")),
            None
        )
        if name_col:
            df["Is a Company"] = df[name_col].apply(infer_is_company)
            logger.info("Inferred 'Is a Company' from '%s' column", name_col)

    # --- state_id / country_id: strip Odoo export artifacts (value = "False") ---
    for col in df.columns:
        if col.lower() in ("state_id", "state", "country_id", "country"):
            df[col] = df[col].apply(
                lambda x: None if (isinstance(x, str) and x.strip().lower() in ("false", "true", "0")) else x
            )

    # --- state / country: scan the ENTIRE ROW for location clues ---
    # Country (and state) can appear anywhere in a row — in the street field,
    # in a city field, or even as the sole value in an address column.
    # e.g. street = "Hubei Province, Baoding City, China"
    #      street = "India"  (entire value is just the country name)
    #      street = "40 KAYODE STREET OGBA IKEJA LAGOS STATE Nigeria"
    # We scan ALL address-like columns on each row, concatenate their text,
    # and run the extraction once on the full combined string.
    country_col = next((c for c in df.columns if c.lower() in ("country_id", "country")), None)
    state_col   = next((c for c in df.columns if c.lower() in ("state_id", "state")), None)

    # Columns that might carry address text in any client file
    _ADDR_HINTS = {"street", "street2", "address", "addr", "city", "area",
                   "location", "region", "zone", "district", "state_id",
                   "country_id", "child_ids/street"}
    addr_cols = [c for c in df.columns
                 if (c.lower() in _ADDR_HINTS
                     or any(k in c.lower() for k in ("street", "address", "addr", "city")))
                 and "type" not in c.lower()    # exclude "Address Type" column
                 and "address type" not in c.lower()]

    for idx, row in df.iterrows():
        state_missing   = (state_col   is not None and
                           (pd.isna(row.get(state_col))   or
                            str(row.get(state_col,   "")).strip() in ("", "None")))
        country_missing = (country_col is not None and
                           (pd.isna(row.get(country_col)) or
                            str(row.get(country_col, "")).strip() in ("", "None")))

        if not (state_missing or country_missing):
            continue

        # Build one combined text string from every address-bearing column in this row
        parts = []
        for col in addr_cols:
            v = row.get(col, "")
            if isinstance(v, str) and v.strip() and v.strip().lower() not in ("false", "none", "nan"):
                parts.append(v.strip())
        combined = " | ".join(parts)   # separator keeps city tokens distinct

        if not combined.strip():
            continue

        state_found, country_found = _extract_state_country_from_row(combined)

        if state_missing and state_found and state_col:
            df.at[idx, state_col] = state_found
        if country_missing and country_found and country_col:
            df.at[idx, country_col] = country_found

    # --- Reference: strip Odoo internal export prefix ---
    ref_cols = [c for c in df.columns if c.lower() in ("reference", "ref")]
    for col in ref_cols:
        df[col] = df[col].apply(_strip_odoo_prefix)

    # --- Credit Limit / Rank: numeric ---
    for col in df.columns:
        schema_key = col  # already renamed by this point
        if schema_key in schema and schema[schema_key]["dtype"] in ("float", "int"):
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(
                schema[schema_key]["default"] or 0
            )

    # --- Zip: force string (avoid float '110001.0') ---
    zip_cols = [c for c in df.columns if c.lower() in ("zip", "postal", "zip code")]
    for col in zip_cols:
        df[col] = df[col].apply(
            lambda x: str(int(float(x))) if (isinstance(x, str) and re.fullmatch(r"\d+\.0", x)) else x
        )

    # --- Account Receivable: derive from Category (FZR-style files) ---
    # This is a fixed business rule observed in your cleaned FZR data:
    #   Horeca       → 112002 Horeca Receivables
    #   Retail       → 112003 Retail Receivables
    #   Distributor  → 112001 Distributors Receivables
    #   Others       → 112005 Other Receivables
    #   Not Sales    → 112006 Related Party Receivables
    # Only fills blank cells — never overwrites a value already present.
    _AR_MAP = {
        "horeca":      "112002 Horeca Receivables",
        "retail":      "112003 Retail Receivables",
        "distributor": "112001 Distributors Receivables",
        "others":      "112005 Other Receivables",
        "not sales":   "112006 Related Party Receivables",
    }
    cat_col = next((c for c in df.columns if c.lower() == "category"), None)
    ar_col  = next((c for c in df.columns if "account receivable" in c.lower()), None)
    if cat_col and data_type == "customer":
        if ar_col is None:
            df["Account Receivable"] = None
            ar_col = "Account Receivable"
        mask = df[ar_col].isna() | (df[ar_col].astype(str).str.strip() == "")
        df.loc[mask, ar_col] = df.loc[mask, cat_col].apply(
            lambda x: _AR_MAP.get(str(x).strip().lower()) if isinstance(x, str) else None
        )

    # --- Note: derive from Branch column when Note is absent ---
    # In FZR raw data, "Branch" contains values like "Office Lagos", "Dave : SP"
    # which your BAs copy into the Note field.
    branch_col = next((c for c in df.columns if c.lower() == "branch"), None)
    note_col   = next((c for c in df.columns if c.lower() == "note"), None)
    if branch_col and data_type == "customer":
        if note_col is None:
            df["Note"] = None
            note_col = "Note"
        mask = df[note_col].isna() | (df[note_col].astype(str).str.strip() == "")
        df.loc[mask, note_col] = df.loc[mask, branch_col]

    return df


# -- helper functions --------------------------------------------------------

def _title_case(val: Any) -> Any:
    if not isinstance(val, str) or not val.strip():
        return val
    return val.strip().title()


def _clean_phone(val: Any) -> Any:
    if not isinstance(val, str) or not val.strip():
        return val
    stripped = val.strip()
    # Preserve leading + for international format, then remove all non-digits
    has_plus = stripped.startswith("+")
    digits = re.sub(r"\D", "", stripped)
    return ("+" if has_plus else "") + digits if digits else None


def _clean_email(val: Any) -> Any:
    if not isinstance(val, str) or not val.strip():
        return val
    val = val.strip().lower()
    match = _EMAIL_RE.search(val)
    return match.group(0) if match else None


def _coerce_bool(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in ("true", "1", "yes", "y")
    return True


# Corporate keywords that strongly indicate a business entity
_COMPANY_KEYWORDS = {
    "ltd", "limited", "co.", "co,", "plc", "inc", "corp", "corporation",
    "llc", "llp", "nig", "nigeria", "ent", "enterprises", "enterprise",
    "int", "international", "services", "service", "solutions", "solution",
    "group", "associates", "association", "holdings", "holding",
    "industries", "industry", "production", "productions", "manufacturing",
    "trading", "logistics", "consulting", "consultancy", "technologies",
    "technology", "systems", "global", "agency", "agencies",
    "pharmaceuticals", "pharmaceutical", "pharmacy", "healthcare",
    "hospital", "clinic", "medical", "surgical", "laboratory", "labs",
    "publishing", "publishers", "media", "communications", "foundation",
    "restaurant", "supermarket", "stores", "store", "market", "bakery",
    "hotel", "suites", "ventures", "investment", "investments", "capital",
    "properties", "property", "real estate", "construction", "contractors",
    "supplies", "supply", "distribution", "distributors", "distributor",
    "imports", "exports", "resources", "management", "academy", "school",
    "college", "university", "institute", "church", "ministry",
    "agro", "agriculture", "agricultural", "produce", "farms", "farm",
    "integrated", "nig.", "and", "sons", "son", "brothers", "bro",
}

# Internal/walk-in account patterns that are NOT real companies
_INTERNAL_ACCOUNT_PATTERNS = re.compile(
    r"walk[\s\-]?in|walk in|cash sale|a\/c\b|debt recovery|"
    r"online customer|regulatory|nafdac sample|^sample$",
    re.IGNORECASE,
)


def infer_is_company(name: Any) -> bool:
    """
    Infer whether a name represents a company or an individual.

    Logic (mirrors what your BAs do manually):
    1. If name contains known corporate suffix/keyword → True
    2. If name matches internal account patterns → False
    3. If name is a single word (no spaces) and short → likely individual → False
    4. If name looks like "Firstname Lastname" (2 words, both capitalised,
       no corporate keywords) → False
    5. Default → True (when in doubt, treat as company)
    """
    if not isinstance(name, str) or not name.strip():
        return True

    name_clean = name.strip()
    name_lower = name_clean.lower()
    words = name_clean.split()

    # Rule 1: corporate keyword anywhere in the name → Company
    name_tokens = set(re.split(r"[\s,.\-\/&()]+", name_lower))
    if name_tokens & _COMPANY_KEYWORDS:
        return True

    # Rule 2: internal account patterns → Individual / non-company
    if _INTERNAL_ACCOUNT_PATTERNS.search(name_lower):
        return False

    # Rule 3: single word (e.g. "Adaora", "Chioma", "George") → Individual
    if len(words) == 1:
        return False

    # Rule 4: exactly 2–3 words, all look like personal names
    # (each word is alphabetic, title-cased, no corporate keyword)
    if len(words) <= 3:
        looks_personal = all(
            re.fullmatch(r"[A-Za-z'\-]+", w) and len(w) >= 2
            for w in words
        )
        if looks_personal:
            return False

    # Rule 5: default → treat as company
    return True


# ---------------------------------------------------------------------------
# Nigerian state extraction from raw street strings
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# World country lookup  –  covers every country seen in your client files
# plus the most common ones a future client might send.
# Key = lowercase (for matching), Value = Odoo-expected proper name.
# Sorted longest-first at build time so multi-word names match before substrings.
# ---------------------------------------------------------------------------
_WORLD_COUNTRIES: dict[str, str] = {
    # Africa
    "nigeria": "Nigeria", "ghana": "Ghana", "kenya": "Kenya",
    "south africa": "South Africa", "ethiopia": "Ethiopia",
    "tanzania": "Tanzania", "uganda": "Uganda", "egypt": "Egypt",
    "cameroon": "Cameroon", "ivory coast": "Ivory Coast",
    "cote d'ivoire": "Ivory Coast", "senegal": "Senegal", "mali": "Mali",
    "angola": "Angola", "mozambique": "Mozambique", "zambia": "Zambia",
    "zimbabwe": "Zimbabwe", "rwanda": "Rwanda", "botswana": "Botswana",
    "namibia": "Namibia", "benin": "Benin", "togo": "Togo",
    "niger": "Niger Republic", "chad": "Chad", "sudan": "Sudan",
    "somalia": "Somalia", "libya": "Libya", "morocco": "Morocco",
    "algeria": "Algeria", "tunisia": "Tunisia",
    # Europe
    "united kingdom": "United Kingdom", "england": "United Kingdom",
    "scotland": "United Kingdom", "wales": "United Kingdom",
    "germany": "Germany", "france": "France", "italy": "Italy",
    "spain": "Spain", "portugal": "Portugal", "netherlands": "Netherlands",
    "belgium": "Belgium", "switzerland": "Switzerland", "austria": "Austria",
    "sweden": "Sweden", "norway": "Norway", "denmark": "Denmark",
    "finland": "Finland", "poland": "Poland", "czech republic": "Czech Republic",
    "hungary": "Hungary", "romania": "Romania", "greece": "Greece",
    "turkey": "Turkey", "russia": "Russia", "ukraine": "Ukraine",
    "ireland": "Ireland",
    # Americas
    "united states": "United States", "usa": "United States",
    "u.s.a": "United States", "u.s.a.": "United States",
    "canada": "Canada", "mexico": "Mexico", "brazil": "Brazil",
    "argentina": "Argentina", "colombia": "Colombia", "chile": "Chile",
    "peru": "Peru", "venezuela": "Venezuela",
    # Asia
    "china": "China", "india": "India", "japan": "Japan",
    "south korea": "South Korea", "indonesia": "Indonesia",
    "pakistan": "Pakistan", "bangladesh": "Bangladesh",
    "vietnam": "Vietnam", "thailand": "Thailand", "malaysia": "Malaysia",
    "singapore": "Singapore", "philippines": "Philippines",
    "hong kong": "Hong Kong", "taiwan": "Taiwan",
    "saudi arabia": "Saudi Arabia", "uae": "United Arab Emirates",
    "united arab emirates": "United Arab Emirates", "dubai": "United Arab Emirates",
    "israel": "Israel", "iran": "Iran", "iraq": "Iraq",
    "jordan": "Jordan", "lebanon": "Lebanon", "qatar": "Qatar",
    "kuwait": "Kuwait", "bahrain": "Bahrain",
    # Oceania
    "australia": "Australia", "new zealand": "New Zealand",
}

# Pre-sort longest-first so "United Kingdom" matches before "United"
_WORLD_COUNTRIES_SORTED: list[tuple[str, str]] = sorted(
    _WORLD_COUNTRIES.items(), key=lambda x: len(x[0]), reverse=True
)

# All 36 states + FCT, normalised to the exact spelling Odoo expects.
# Built from every state value seen across all your cleaned client files.
_NG_STATES: dict[str, str] = {
    "abia": "Abia", "adamawa": "Adamawa", "akwa ibom": "Akwa Ibom",
    "akwaibom": "Akwa Ibom", "anambra": "Anambra", "bauchi": "Bauchi",
    "bayelsa": "Bayelsa", "benue": "Benue", "borno": "Borno",
    "cross river": "Cross River", "crossriver": "Cross River",
    "delta": "Delta", "ebonyi": "Ebonyi", "edo": "Edo",
    "ekiti": "Ekiti", "enugu": "Enugu", "fct": "FCT", "abuja": "FCT",
    "gombe": "Gombe", "imo": "Imo", "jigawa": "Jigawa",
    "kaduna": "Kaduna", "kano": "Kano", "katsina": "Katsina",
    "kebbi": "Kebbi", "kogi": "Kogi", "kwara": "Kwara",
    "lagos": "Lagos", "nasarawa": "Nasarawa", "niger": "Niger",
    "ogun": "Ogun", "ondo": "Ondo", "osun": "Osun", "oyo": "Oyo",
    "plateau": "Plateau", "plateau state": "Plateau",
    "rivers": "Rivers", "sokoto": "Sokoto", "taraba": "Taraba",
    "yobe": "Yobe", "zamfara": "Zamfara",
}

# City → State lookup: derived from every city/state pair in your cleaned files.
# Covers ~200+ Nigerian cities/areas. Add more rows freely as new clients appear.
_CITY_TO_STATE: dict[str, str] = {
    "aba": "Abia", "abakaliki": "Ebonyi", "abeokuta": "Ogun",
    "abuja": "FCT", "abulado": "Lagos", "abule ado": "Lagos",
    "abule egba": "Lagos", "abule iroko": "Ogun", "abule odu": "Lagos",
    "ado ekiti": "Ekiti", "ado-ekiti": "Ekiti", "agboju": "Lagos",
    "agege": "Lagos", "agidingbi": "Lagos", "aguda": "Lagos",
    "ajah": "Lagos", "ajao estate": "Lagos", "ajegunle": "Lagos",
    "akoka": "Lagos", "akure": "Ondo", "akute": "Ogun",
    "akwa": "Anambra", "alagbado": "Lagos", "alausa": "Lagos",
    "allen": "Lagos", "amuwo odofin": "Lagos", "amuwo-odofin": "Lagos",
    "anthony": "Lagos", "anthony village": "Lagos", "apapa": "Lagos",
    "apete": "Oyo", "apo": "FCT", "apongbon": "Lagos",
    "arepo": "Ogun", "asaba": "Delta", "asokoro": "FCT",
    "auchi": "Edo", "awkunanaw": "Enugu", "awoyaya": "Lagos",
    "badagry": "Lagos", "badagery": "Lagos", "badagary": "Lagos",
    "bariga": "Lagos", "bauchi": "Bauchi", "benin": "Edo",
    "benin city": "Edo", "berger": "Lagos", "bori": "Rivers",
    "calabar": "Cross River", "dopemu": "Lagos",
    "duboyi": "FCT", "dutse": "FCT", "dutse apo": "FCT",
    "ebute meta": "Lagos", "ebute metta": "Lagos", "ebute-metta": "Lagos",
    "egbe ikotun": "Lagos", "egbeda": "Lagos", "ejigbo": "Lagos",
    "eket": "Akwa Ibom", "elelenwo": "Rivers", "enugu": "Enugu",
    "eruwa": "Oyo", "fadeyi": "Lagos", "fastac town": "Lagos",
    "festac": "Lagos", "festac town": "Lagos", "festac/amuwo": "Lagos",
    "galadimawa": "FCT", "garki": "FCT", "gbagada": "Lagos",
    "gombe": "Gombe", "gudu": "FCT", "gwagwalada": "FCT",
    "gwarimpa": "FCT", "gwaska": "FCT", "ibadan": "Oyo",
    "ibafo": "Lagos", "ibeju lekki": "Lagos", "ibeju-lekki": "Lagos",
    "idimu": "Lagos", "idu": "FCT", "idumota": "Lagos",
    "ifako": "Lagos", "ifako-agege": "Lagos", "igando": "Lagos",
    "iganmu": "Lagos", "ijebu-ode": "Ogun", "ijebu ode": "Ogun",
    "ijebu mushin": "Ogun", "ijegun": "Lagos", "ijeodo": "Lagos",
    "ijeshatedo": "Lagos", "ijora": "Lagos", "iju": "Lagos",
    "iju ishaga": "Lagos", "ikate": "Lagos", "ikeja": "Lagos",
    "ikorodu": "Lagos", "ikosi ketu": "Lagos", "ikosi-ketu": "Lagos",
    "ikota": "Lagos", "ikotun": "Lagos", "ikoyi": "Lagos",
    "ilaro": "Ogun", "ilasa": "Lagos", "ilishan-remo": "Ogun",
    "ilorin": "Kwara", "ilupeju": "Lagos", "illupeju": "Lagos",
    "ipaja": "Lagos", "isale-eko": "Lagos", "isale eko": "Lagos",
    "iseyin": "Oyo", "ishaga": "Lagos", "isheri": "Lagos",
    "isolo": "Lagos", "itafaji": "Lagos", "itire": "Lagos",
    "iyana ipaja": "Lagos", "jabi": "FCT", "jahi": "FCT",
    "jericho": "Oyo", "jibowu": "Lagos", "jos": "Plateau",
    "kado": "FCT", "kaduna": "Kaduna", "kano": "Kano",
    "karu": "FCT", "katampe": "FCT", "ketu": "Lagos",
    "kubwa": "FCT", "kuje": "FCT", "lafiaji": "Lagos",
    "lagos": "Lagos", "lagos island": "Lagos", "lekki": "Lagos",
    "lekki ajah": "Lagos", "lekki phase 1": "Lagos", "lekki vgc": "Lagos",
    "life camp": "FCT", "lokoja": "Kogi", "lugbe": "FCT",
    "mabushi": "FCT", "magodo": "Lagos", "maitama": "FCT",
    "makurdi": "Benue", "mararaba": "Nasarawa", "marina": "Lagos",
    "maroko": "Lagos", "maryland": "Lagos", "matori": "Lagos",
    "mende": "Lagos", "mile 2": "Lagos", "mowe": "Ogun",
    "mowe/ibafo": "Ogun", "mushin": "Lagos", "nasarawa": "Nasarawa",
    "nnewi": "Anambra", "nyanya": "FCT", "obanikoro": "Lagos",
    "obantoko": "Ogun", "ogba": "Lagos", "ogbomosho": "Oyo",
    "ogere": "Ogun", "ogijo": "Ogun", "ogudu": "Lagos",
    "ojo": "Lagos", "ojodu": "Lagos", "ojodu berger": "Lagos",
    "ojokoro": "Lagos", "ojota": "Lagos", "okota": "Lagos",
    "okitipupa": "Ondo", "oko oba": "Lagos", "olodi": "Lagos",
    "onigbongbo": "Lagos", "onikan": "Lagos", "onipanu": "Lagos",
    "oniru": "Lagos", "onitsha": "Anambra", "oregun": "Lagos",
    "orile iganmu": "Lagos", "osborne": "Lagos", "oshodi": "Lagos",
    "oshodi-isolo": "Lagos", "osogbo": "Osun", "ota": "Ogun",
    "owerri": "Imo", "owerrinta": "Abia", "owo": "Ondo",
    "oworonshoki": "Lagos", "oworonsoki": "Lagos",
    "palmgrove": "Lagos", "pedro": "Lagos", "port harcourt": "Rivers",
    "ph": "Rivers", "sabo": "Lagos", "sagamu": "Ogun",
    "saminaka": "Kaduna", "sango ota": "Ogun", "sango-ota": "Ogun",
    "sangotedo": "Lagos", "satellite town": "Lagos",
    "satalite town": "Lagos", "satelite town": "Lagos",
    "shangisha": "Lagos", "shomolu": "Lagos", "somolu": "Lagos",
    "suleja": "Niger", "surulere": "Lagos", "tanke": "Kwara",
    "tinubu square": "Lagos", "ughelli": "Delta", "ugheli": "Delta",
    "umuahia": "Abia", "umunze": "Anambra", "uwani": "Enugu",
    "uyo": "Akwa Ibom", "vgc": "Lagos", "victoria island": "Lagos",
    "warri": "Delta", "woji": "Rivers", "wuse": "FCT",
    "wuse 2": "FCT", "wuse zone 5": "FCT", "yaba": "Lagos",
    "yenagoa": "Bayelsa", "yenegoa": "Bayelsa",
    "zone 5": "FCT", "zone 6 abuja": "FCT",
}

# ---------------------------------------------------------------------------
# Learned city cache  –  grows automatically as the AI resolves new cities
# ---------------------------------------------------------------------------
# The static _CITY_TO_STATE table above is a fast first-pass lookup for known
# Nigerian cities. But new clients will always bring cities not in that table.
# Every time the AI resolves a city → state mapping that isn't already known,
# we write it to a JSON cache file next to this script. On the next run the
# engine loads that cache and merges it into _CITY_TO_STATE — so the table
# gets smarter with every file processed, with zero manual maintenance.

_CITY_CACHE_PATH = Path(__file__).parent / "learned_cities.json"


def _load_city_cache() -> dict[str, str]:
    """Load previously learned city→state mappings from disk."""
    if _CITY_CACHE_PATH.exists():
        try:
            with _CITY_CACHE_PATH.open("r", encoding="utf-8") as fh:
                cache = json.load(fh)
            logger.info("Loaded %d learned city mappings from cache", len(cache))
            return {k.lower(): v for k, v in cache.items()}
        except Exception as exc:
            logger.warning("Could not load city cache: %s", exc)
    return {}


def _save_city_cache(cache: dict[str, str]) -> None:
    """Persist the city→state cache to disk."""
    try:
        with _CITY_CACHE_PATH.open("w", encoding="utf-8") as fh:
            json.dump(dict(sorted(cache.items())), fh, indent=2, ensure_ascii=False)
    except Exception as exc:
        logger.warning("Could not save city cache: %s", exc)


# Merge learned cache into the live lookup table at import time
_CITY_TO_STATE.update(_load_city_cache())

# Build a single flat regex that matches any state name at a word boundary.
# Sorted longest-first so "Akwa Ibom" matches before "Akwa".
_STATE_NAMES_SORTED = sorted(_NG_STATES.keys(), key=len, reverse=True)
_STATE_REGEX = re.compile(
    r"\b(" + "|".join(re.escape(s) for s in _STATE_NAMES_SORTED) + r")\s*(state)?\b",
    re.IGNORECASE,
)


def _extract_state_country_from_row(combined: str) -> tuple[str | None, str | None]:
    """
    Given a combined text string built from ALL address columns in a row,
    return (state, country).  Works for any country — not just Nigeria.

    Strategy (pure Python, no LLM, cheapest-first):

    Pass 0 — Explicit country name anywhere in the text
        Checks the combined text against the world country table.
        Handles cases like street = "India", or "...CHINA" at end of string.

    Pass 1 — Nigerian state name match (skipping "X Street/Road" false matches)
        If a Nigerian state is found → country = Nigeria.

    Pass 2 — City → Nigerian state lookup on every token window
        Catches cases where only a city name is present (e.g. "Ikeja", "Lekki").

    Pass 3 — Nothing found → return (None, None)
        The AI address parser will handle it in the next stage if needed.
    """
    if not isinstance(combined, str) or not combined.strip():
        return None, None

    text_lower = combined.strip().lower()

    # --- Pass 0: explicit world country name anywhere in text ---
    # Sorted longest-first to match "United Kingdom" before "United"
    # If a territory (Hong Kong) AND its sovereign country (China) both appear,
    # prefer the sovereign country — scan all matches, return the non-territory one.
    _TERRITORY_NAMES = {"hong kong", "macau", "macao", "puerto rico", "dubai", "taiwan"}
    found_countries: list[str] = []
    for country_lower, country_proper in _WORLD_COUNTRIES_SORTED:
        if re.search(r"\b" + re.escape(country_lower) + r"\b", text_lower):
            if country_lower == "nigeria":
                break   # fall through to state extraction
            found_countries.append((country_lower, country_proper))
    if found_countries:
        # Prefer sovereign over territory when both present
        sovereign = [(cl, cp) for cl, cp in found_countries if cl not in _TERRITORY_NAMES]
        best = sovereign[0] if sovereign else found_countries[0]
        return None, best[1]

    # --- Pass 1: Nigerian state name match ---
    _STREET_WORDS_RE = re.compile(
        r"^(street|road|avenue|close|lane|way|crescent|drive|place|court|str|rd|blvd)\b",
        re.IGNORECASE,
    )
    for m in _STATE_REGEX.finditer(text_lower):
        raw_match = m.group(1).lower()
        state = _NG_STATES.get(raw_match)
        if not state:
            continue
        after = text_lower[m.end():].strip().lstrip(",. |")
        if _STREET_WORDS_RE.match(after):
            continue   # "Kano Street" is not Kano state
        # Also skip if followed by "word + expressway/road/street"
        # e.g. "LAGOS IBADAN EXPRESSWAY" — Lagos here is part of a road name
        _COMPOUND_ROAD_RE = re.compile(
            r"^\w+\s+(expressway|highway|bypass|motorway|freeway|road|street|way)\b",
            re.IGNORECASE,
        )
        if _COMPOUND_ROAD_RE.match(after):
            continue
        return state, "Nigeria"

    # --- Pass 2: city → Nigerian state lookup ---
    _STREET_SUFFIXES = {"street", "road", "avenue", "close", "lane", "way", "crescent",
                        "drive", "place", "court", "boulevard", "expressway", "bypass",
                        "highway", "freeway", "motorway", "str", "rd", "express"}
    tokens = re.split(r"[,|\s]+", text_lower.rstrip("."))
    tokens = [t for t in tokens if t]

    # Try trailing tokens first (country/state usually at the end)
    for window in (3, 2, 1):
        if len(tokens) >= window:
            candidate = " ".join(tokens[-window:])
            state = _CITY_TO_STATE.get(candidate)
            if state:
                return state, "Nigeria"

    # Then full scan, skipping street-name tokens
    for i in range(len(tokens)):
        if i + 1 < len(tokens) and tokens[i + 1] in _STREET_SUFFIXES:
            continue
        for window in (3, 2, 1):
            phrase = " ".join(tokens[i : i + window])
            state = _CITY_TO_STATE.get(phrase)
            if state:
                return state, "Nigeria"

    return None, None


def _strip_odoo_prefix(val: Any) -> Any:
    """Remove Odoo internal export IDs like '__export__.res_partner_1234_abc' → keep the ref."""
    if not isinstance(val, str):
        return val
    if val.startswith("__export__"):
        return None    # discard internal IDs entirely – not useful in imports
    return val.strip()


# ===========================================================================
# 5. AI ADDRESS PARSER  –  batch LLM call for unstructured address strings
# ===========================================================================

def ai_parse_addresses(
    addresses: list[str],
    client: anthropic.Anthropic,
    batch_size: int = 20,
) -> list[dict]:
    """
    Given a list of raw address strings, return a list of dicts with keys:
    street, street2, city, state, zip, country.

    Batches calls to keep cost low and avoid timeouts.
    """
    results: list[dict] = []
    empty = {"street": None, "street2": None, "city": None,
             "state": None, "zip": None, "country": None}

    for i in range(0, len(addresses), batch_size):
        batch = addresses[i : i + batch_size]
        parsed = _ai_address_batch(batch, client)
        results.extend(parsed)
        if i + batch_size < len(addresses):
            time.sleep(0.3)   # light rate-limit courtesy pause

    # Safety: ensure list length matches input
    while len(results) < len(addresses):
        results.append(empty.copy())
    return results


def _ai_address_batch(batch: list[str], client: anthropic.Anthropic) -> list[dict]:
    empty = {"street": None, "street2": None, "city": None,
             "state": None, "zip": None, "country": None}

    numbered = {str(i): addr for i, addr in enumerate(batch)}

    system_prompt = (
        "You are an international address parsing specialist. "
        "Return ONLY valid JSON – no prose, no code fences."
    )
    user_prompt = f"""Parse each address string into clean, structured Odoo address components.
These strings may be messy: combined fields, extra tokens, inconsistent formatting.

INPUT (JSON object, key = index, value = raw combined address string):
{json.dumps(numbered, indent=2)}

Return a JSON object where each key is the same index and the value is:
{{
  "street":  "house number + street name only",
  "street2": "estate / area / landmark / additional direction, or null",
  "city":    "city or town name only",
  "state":   "state, county, or province name only",
  "zip":     "postal or zip code, or null",
  "country": "full sovereign country name, or null if genuinely ambiguous"
}}

PARSING RULES — read carefully:
- Extract ONLY what belongs in each field. Do not bleed one field into another.
- "street"  = the actual house/plot number + street name. No city, no state, no country.
- "street2" = estate name, area name, landmark, floor/suite — only if present.
- "city"    = the city or town. NOT the state. Lagos is both a city and a state —
              if the address says "Ikeja, Lagos", then city=Ikeja, state=Lagos.
              If only "Lagos" appears with no other city clue, city=Lagos, state=Lagos.
- "state"   = Nigerian state (Lagos, Ogun, Rivers, FCT, Kano, etc.) or equivalent
              for other countries. NEVER put a city name here.
- "country" = infer from city/state/street context. Default to "Nigeria" only when
              all other signals point to Nigeria. Use null if truly ambiguous.
- Title-case all output values.
- Use null (never empty string) for missing fields.

Return ONLY the JSON object. No prose, no markdown.
"""

    max_retries = 3
    for attempt in range(max_retries):
        try:
            raw_text = _complete_with_fallback(client, system_prompt, user_prompt, max_tokens=1500)
            raw_text = re.sub(r"^```[a-z]*\n?|```$", "", raw_text, flags=re.MULTILINE).strip()
            parsed_map: dict[str, dict] = json.loads(raw_text)
            return [parsed_map.get(str(i), empty.copy()) for i in range(len(batch))]
        except (json.JSONDecodeError, _LLMUnavailable) as exc:
            logger.warning("Address batch attempt %d failed: %s", attempt + 1, exc)
            time.sleep(2 ** attempt)

    return [empty.copy() for _ in batch]


# ===========================================================================
# 5b. AI COUNTRY / STATE RESOLVER  –  fallback for rows rule-based can't resolve
# ===========================================================================

def ai_resolve_country_state(
    rows: list[dict],          # list of {idx, combined_address} dicts
    client: anthropic.Anthropic,
    batch_size: int = 30,
) -> dict[int, dict]:
    """
    For rows where rule-based extraction found neither country nor state,
    ask the AI to infer them from whatever address text exists in the row.

    Only called for rows that actually need it — keeps API cost minimal.
    Returns a dict: { original_row_idx: {"state": ..., "country": ...} }
    """
    results: dict[int, dict] = {}
    empty = {"state": None, "country": None}

    for i in range(0, len(rows), batch_size):
        batch = rows[i : i + batch_size]
        batch_results = _ai_country_state_batch(batch, client)
        results.update(batch_results)
        if i + batch_size < len(rows):
            time.sleep(0.3)

    return results


def _ai_country_state_batch(
    rows: list[dict],
    client: anthropic.Anthropic,
) -> dict[int, dict]:
    """
    Single batched LLM call to resolve country and state from raw address text.
    Input rows: [{"idx": int, "combined_address": str}, ...]
    Returns:    {idx: {"state": str|None, "country": str|None}}
    """
    empty = {"state": None, "country": None}
    numbered = {str(r["idx"]): r["combined_address"] for r in rows}

    system_prompt = (
        "You are an international address parsing specialist. "
        "Return ONLY valid JSON – no prose, no code fences."
    )
    user_prompt = f"""For each address string below, identify the country, state/region, and city.
These addresses may be partial, misspelled, or have no space between city and country.

INPUT (JSON object — key = row index, value = raw address text):
{json.dumps(numbered, indent=2)}

Return a JSON object where each key is the same row index and the value is:
{{"state": "...", "country": "...", "city": "..."}}

Rules:
- "country" = full sovereign country name (e.g. "Nigeria", "United Kingdom", "China", "India").
  Infer from city names, street names, postal codes, language patterns, or any other clue.
  Examples: "AJEGUNLE" → Nigeria, "Chichester" → United Kingdom, "New Delhi" → India,
  "Portharcourt" or "PORTHACOURT" (typo) → Nigeria (Rivers State).
- "state" = state, county, province, or region within that country.
  For Nigeria use exact state names: Lagos, Rivers, Ogun, FCT, Kano, etc.
  For UK use county names. For US use state names. For others use province/region.
- "city" = the canonical, correctly-spelled city or town name you identified.
  This is used to build a lookup cache so the same city never needs AI resolution again.
  Use null if no city is identifiable.
- Use null for state/country/city if the address is completely ambiguous (e.g. just a P.O. Box).
- Title-case all values.
- Do NOT default to Nigeria — only use it when you are confident.

Return ONLY the JSON object.
"""

    max_retries = 3
    for attempt in range(max_retries):
        try:
            raw_text = _complete_with_fallback(client, system_prompt, user_prompt, max_tokens=1000)
            raw_text = re.sub(r"^```[a-z]*\n?|```$", "", raw_text, flags=re.MULTILINE).strip()
            parsed: dict[str, dict] = json.loads(raw_text)
            return {r["idx"]: parsed.get(str(r["idx"]), empty.copy()) for r in rows}
        except (json.JSONDecodeError, _LLMUnavailable) as exc:
            logger.warning("AI country/state batch attempt %d failed: %s", attempt + 1, exc)
            time.sleep(2 ** attempt)

    return {r["idx"]: empty.copy() for r in rows}


# ===========================================================================
# 5c. AI FIELD-LEVEL CLEANER  –  fixes cross-contaminated fields
# ===========================================================================

def ai_clean_flagged_fields(
    df: pd.DataFrame,
    flagged_fields: list[str],
    client: anthropic.Anthropic,
    batch_size: int = 25,
) -> pd.DataFrame:
    """
    For fields the column mapper flagged as containing wrong data
    (e.g. phone col has person names, state col has city names),
    send each unique dirty value to the AI in one batched call and
    get back a clean corrected value — or null if the value does
    not belong in that field at all.

    Examples:
      Phone = "PHARM GODWIN OSUJI"         → null  (it's a name, not a phone)
      Phone = "08095000005,  08033060854"  → "08095000005"  (keep first)
      Mobile = "08030511797 Kabiru Abubakar" → "08030511797" (strip the name)
      State = "Ikeja"                      → null  (it's a city, not a state)
      Country = "08012345678"              → null  (it's a phone, not a country)
    """
    for field in flagged_fields:
        if field not in df.columns:
            continue

        # Only process non-null values — collect unique dirty ones to minimise tokens
        mask_dirty = df[field].notna() & (df[field].astype(str).str.strip() != "")
        unique_vals = df.loc[mask_dirty, field].astype(str).unique().tolist()

        if not unique_vals:
            continue

        logger.info("AI field clean: '%s' — %d unique values to check", field, len(unique_vals))

        # Batch the unique values
        corrections: dict[str, str | None] = {}
        for i in range(0, len(unique_vals), batch_size):
            batch = unique_vals[i : i + batch_size]
            batch_result = _ai_field_clean_batch(field, batch, client)
            corrections.update(batch_result)
            if i + batch_size < len(unique_vals):
                time.sleep(0.3)

        # Apply corrections back to DataFrame
        df[field] = df[field].astype(str).map(
            lambda v: corrections.get(v, v) if pd.notna(v) and v.strip() else v
        )
        # Treat "None" string results as actual None
        df[field] = df[field].replace({"None": None, "nan": None, "": None})

    return df


def _ai_field_clean_batch(
    field_name: str,
    values: list[str],
    client: anthropic.Anthropic,
) -> dict[str, str | None]:
    """
    Single batched LLM call to clean a list of potentially-wrong values
    for a specific Odoo field.
    """
    numbered = {str(i): v for i, v in enumerate(values)}

    # Field-specific cleaning rules — tailored instructions per field type
    # produce much more precise AI output than generic rules
    _field_rules: dict[str, str] = {
        "phone": (
            "Strip any person name or label prefix. Keep only the phone number digits and "
            "leading +. If the value contains no digits at all, return null."
        ),
        "mobile": (
            "Strip any person name or label prefix. Keep only the phone number digits and "
            "leading +. If multiple numbers appear, keep the first. If no digits, return null."
        ),
        "state": (
            "The value should be a Nigerian state name (e.g. Lagos, Kano, Rivers, FCT). "
            "If it is a city name or any non-state value, return null. "
            "Correct obvious typos (e.g. 'Lagod' → 'Lagos'). Return the state name only."
        ),
        "country": (
            "The value should be a country name only (e.g. Nigeria, Ghana, United Kingdom). "
            "If it contains a full address, city, or is not a country name, return null. "
            "Normalise to the standard English country name."
        ),
    }
    rule = _field_rules.get(field_name.lower(), "Clean the value — remove noise, fix obvious errors.")

    system_prompt = (
        "You are a data cleaning specialist. "
        "Return ONLY valid JSON — no prose, no markdown, no code fences."
    )
    user_prompt = f"""Clean the following values for the Odoo field "{field_name}".

RULE: {rule}

VALUES TO CLEAN (JSON array):
{json.dumps(values, indent=2)}

Return a JSON object mapping each original value to its cleaned version,
or null if the value should be blanked out entirely:
{{
  "original value 1": "cleaned value 1",
  "original value 2": null
}}

Return ONLY the JSON object.
"""

    max_retries = 3
    for attempt in range(max_retries):
        try:
            raw_text = _complete_with_fallback(client, system_prompt, user_prompt, max_tokens=1000)
            raw_text = re.sub(r"^```[a-z]*\n?|```$", "", raw_text, flags=re.MULTILINE).strip()
            result: dict[str, str | None] = json.loads(raw_text)
            return {str(k): (str(v) if v is not None else None) for k, v in result.items()}
        except (json.JSONDecodeError, _LLMUnavailable) as exc:
            logger.warning("Field clean batch attempt %d failed for '%s': %s", attempt + 1, field_name, exc)
            time.sleep(2 ** attempt)

    logger.warning("Field clean failed for '%s' — returning values unchanged", field_name)
    return {v: v for v in values}


# ===========================================================================
# 6. COLUMN REMAPPER  –  apply AI mapping + fill defaults + reorder
# ===========================================================================

def apply_column_mapping(
    df: pd.DataFrame,
    mapping: dict[str, str | None],
    schema: dict[str, dict],
    data_type: str,
    client: anthropic.Anthropic,
    flags: dict[str, list[str]] | None = None,
) -> pd.DataFrame:
    """
    1. Rename raw columns to Odoo field names per AI mapping.
    2. AI address splitting  — driven by flags["needs_address_split"] from the column
       mapper. No heuristics. The AI already saw the data and flagged what needs splitting.
    3. AI field cleaning     — driven by flags["needs_field_clean"]. Fixes cross-field
       contamination: phone cols with names, state cols with cities, etc.
    4. AI country/state fallback for rows still missing location after steps 2-3.
    5. Drop unmapped columns.
    6. Add missing schema columns with default values.
    7. Reorder to match schema order.
    """
    flags = flags or {"needs_address_split": [], "needs_field_clean": []}
    # --- Step 1: Rename ---
    rename_map: dict[str, str] = {k: v for k, v in mapping.items() if v is not None}
    df = df.rename(columns=rename_map)

    # Drop duplicate columns (keep first occurrence after rename)
    df = df.loc[:, ~df.columns.duplicated(keep="first")]

    # --- Step 2: AI address splitting — only for columns the AI flagged ---
    # The column mapper already saw sample data and flagged raw columns whose values
    # are combined address strings needing splitting. We use that judgement directly
    # instead of a comma-counting heuristic that can't understand context.
    _addr_field_map = {
        "street":  "Street",
        "street2": "Street2",
        "city":    "City",
        "state":   "State",
        "zip":     "Zip",
        "country": "Country",
    }
    # Resolve flagged raw col names → their post-rename Odoo field names
    needs_split_odoo = set()
    for raw_col in flags.get("needs_address_split", []):
        odoo_name = mapping.get(raw_col)
        if odoo_name and odoo_name in df.columns:
            needs_split_odoo.add(odoo_name)
        elif raw_col in df.columns:
            needs_split_odoo.add(raw_col)

    if needs_split_odoo:
        logger.info("AI address splitting flagged columns: %s", sorted(needs_split_odoo))
        # Build one combined address string per row from all flagged columns
        # (handles cases where address is split across two flagged columns)
        combined_addrs = (
            df[sorted(needs_split_odoo)]
            .fillna("")
            .astype(str)
            .apply(lambda row: " | ".join(v.strip() for v in row if v.strip()), axis=1)
            .tolist()
        )
        parsed = ai_parse_addresses(combined_addrs, client)
        addr_df = pd.DataFrame(parsed)

        for ai_key, odoo_field in _addr_field_map.items():
            parsed_vals = addr_df[ai_key].values
            if odoo_field not in df.columns:
                df[odoo_field] = parsed_vals
            else:
                # Only overwrite cells that are blank OR came from a flagged column
                mask = df[odoo_field].isna() | (df[odoo_field].astype(str).str.strip() == "")
                df.loc[mask, odoo_field] = [parsed_vals[i] for i in df.index[mask]]

        # Clear the original flagged columns — their content has been redistributed
        for col in needs_split_odoo:
            if col in df.columns and col not in _addr_field_map.values():
                df.drop(columns=[col], inplace=True, errors="ignore")

    # --- Step 2c: AI field-level cleaning for cross-contaminated fields ---
    # The column mapper flagged Odoo fields whose values contain wrong data
    # (phone with names, state with city values, etc.). We fix those now.
    needs_clean = flags.get("needs_field_clean", [])
    if needs_clean:
        logger.info("AI field cleaning flagged fields: %s", needs_clean)
        df = ai_clean_flagged_fields(df, needs_clean, client)

    # --- Step 2b: AI country/state fallback for unresolved rows ---
    # After rule_based_clean has already tried _extract_state_country_from_row,
    # some rows still have blank Country and State (e.g. partial addresses, typos,
    # addresses where no country/city keyword was recognisable).
    # We collect only those rows, send them in one batched AI call, and fill the gaps.
    # Cost: typically 0–30 rows per file → 0–1 extra API calls.
    _country_col = next((c for c in df.columns if c.lower() in ("country", "country_id")), None)
    _state_col   = next((c for c in df.columns if c.lower() in ("state", "state_id")), None)
    _addr_hint_keys = {"street", "street2", "city", "area", "location", "region",
                       "district", "child_ids/street"}
    _addr_text_cols = [c for c in df.columns
                       if c.lower() in _addr_hint_keys
                       or any(k in c.lower() for k in ("street", "address", "addr", "city"))
                       and "type" not in c.lower()]

    _unresolved: list[dict] = []
    for idx, row in df.iterrows():
        country_blank = (
            _country_col is None or
            pd.isna(row.get(_country_col)) or
            str(row.get(_country_col, "")).strip() in ("", "None")
        )
        state_blank = (
            _state_col is None or
            pd.isna(row.get(_state_col)) or
            str(row.get(_state_col, "")).strip() in ("", "None")
        )
        if not (country_blank or state_blank):
            continue   # already resolved — skip

        # Build combined address text from all address-type columns in this row
        parts = [
            str(row[c]).strip() for c in _addr_text_cols
            if isinstance(row.get(c), str)
            and row[c].strip()
            and row[c].strip().lower() not in ("false", "none", "nan")
        ]
        combined = " | ".join(parts)
        if combined.strip():
            _unresolved.append({"idx": idx, "combined_address": combined})

    if _unresolved:
        logger.info(
            "AI country/state fallback: resolving %d unresolved rows …", len(_unresolved)
        )
        ai_geo = ai_resolve_country_state(_unresolved, client)

        for item in _unresolved:
            idx = item["idx"]
            resolved = ai_geo.get(idx, {})
            ai_state   = resolved.get("state")
            ai_country = resolved.get("country")

            if _country_col and ai_country and (
                pd.isna(df.at[idx, _country_col]) or
                str(df.at[idx, _country_col]).strip() in ("", "None")
            ):
                df.at[idx, _country_col] = ai_country

            if _state_col and ai_state and (
                pd.isna(df.at[idx, _state_col]) or
                str(df.at[idx, _state_col]).strip() in ("", "None")
            ):
                df.at[idx, _state_col] = ai_state

        resolved_count = sum(
            1 for item in _unresolved
            if ai_geo.get(item["idx"], {}).get("country")
        )
        logger.info(
            "AI country/state fallback: resolved %d/%d rows",
            resolved_count, len(_unresolved)
        )

        # --- Cache newly learned city → state mappings ---
        # For every row the AI resolved, if it returned a city name and that
        # city isn't already in our lookup table, add it to the persistent cache
        # so future runs won't need to call the AI for that city again.
        new_entries: dict[str, str] = {}
        for item in _unresolved:
            resolved = ai_geo.get(item["idx"], {})
            ai_city  = resolved.get("city")
            ai_state = resolved.get("state")
            ai_ctry  = resolved.get("country")
            # Only cache Nigerian city→state (that's what our lookup table covers)
            if ai_city and ai_state and ai_ctry == "Nigeria":
                city_key = ai_city.strip().lower()
                if city_key and city_key not in _CITY_TO_STATE:
                    new_entries[city_key] = ai_state
                    _CITY_TO_STATE[city_key] = ai_state   # live update for this run too

        if new_entries:
            # Merge into the on-disk cache
            existing_cache = {}
            if _CITY_CACHE_PATH.exists():
                try:
                    with _CITY_CACHE_PATH.open("r", encoding="utf-8") as fh:
                        existing_cache = json.load(fh)
                except Exception:
                    pass
            existing_cache.update(new_entries)
            _save_city_cache(existing_cache)
            logger.info(
                "City cache: added %d new entries %s",
                len(new_entries), list(new_entries.items())[:5]
            )

    # --- Step 3: Drop unmapped / unknown columns ---
    known_cols = set(schema.keys())
    cols_to_keep = [c for c in df.columns if c in known_cols]
    df = df[cols_to_keep].copy()

    # --- Step 4: Add missing schema columns with defaults ---
    for field, meta in schema.items():
        if field not in df.columns:
            df[field] = meta["default"]

    # --- Step 5: Reorder to schema order ---
    df = df[[f for f in schema.keys() if f in df.columns]]

    return df


# ===========================================================================
# 7. VALIDATOR  –  mandatory field checks + error log
# ===========================================================================

def validate_and_split(
    df: pd.DataFrame,
    data_type: str,
) -> tuple[pd.DataFrame, pd.DataFrame, list[int]]:
    """
    Check every row for the mandatory field, but — unlike the old behavior —
    do NOT remove failing rows from the main DataFrame. They stay in place
    (so the output file's row count always matches the input) and are
    instead flagged for the caller to highlight.

    Returns:
      - df: the same DataFrame, unchanged (all rows, original order)
      - error_df: just the failing rows, with '_errors'/'_source_row' columns
        added, for a separate explanatory sheet
      - error_positions: 0-based *positional* row numbers (not pandas index
        labels — upstream steps can leave index gaps, e.g. from dropping
        blank rows without a reset_index) of the failing rows, so the caller
        can map them to the correct Excel row number for highlighting.
    """
    mandatory = MANDATORY_FIELD[data_type]
    errors: list[dict] = []
    error_positions: list[int] = []

    for pos, (_, row) in enumerate(df.iterrows()):
        row_errors: list[str] = []

        # Check mandatory field
        val = row.get(mandatory)
        if pd.isna(val) or (isinstance(val, str) and not val.strip()):
            row_errors.append(f"Missing mandatory field: '{mandatory}'")

        if row_errors:
            err_row = row.to_dict()
            err_row["_errors"] = " | ".join(row_errors)
            err_row["_source_row"] = pos + 2   # Excel 1-indexed + header row
            errors.append(err_row)
            error_positions.append(pos)

    error_df = pd.DataFrame(errors)

    logger.info(
        "Validation: %d clean rows, %d error rows",
        len(df) - len(error_positions), len(error_df)
    )
    return df, error_df, error_positions


# ===========================================================================
# 8. MAIN PIPELINE ORCHESTRATOR
# ===========================================================================

def process_file(
    input_path: str | Path,
    data_type: str,                    # "customer" or "vendor"
    output_path: str | Path | None = None,
    api_key: str | None = None,
    address_batch_size: int = 20,
) -> dict:
    """
    Full end-to-end pipeline.

    Returns:
        {
            "status": "success" | "partial" | "error",
            "message": str,
            "output_path": str | None,
            "stats": { "total": int, "clean": int, "errors": int }
        }
    """
    input_path = Path(input_path)
    if data_type not in SCHEMA_MAP:
        raise ValueError(f"data_type must be 'customer' or 'vendor', got: {data_type!r}")

    schema = SCHEMA_MAP[data_type]

    # --- Resolve output path ---
    if output_path is None:
        output_path = input_path.parent / f"{input_path.stem}_ODOO_READY.xlsx"
    output_path = Path(output_path)

    # --- Anthropic client ---
    # max_retries=ANTHROPIC_MAX_RETRIES (default 3): the SDK retries a 429
    # internally before raising — giving Claude a few chances to recover from
    # a transient rate limit before _complete_with_fallback() drops to the
    # lower-quality free OpenRouter model. Once those retries are exhausted,
    # the exception reaches our code and the fallback takes over.
    client = anthropic.Anthropic(api_key=api_key, max_retries=ANTHROPIC_MAX_RETRIES)  # api_key=None → reads ANTHROPIC_API_KEY env var

    try:
        # 1. Load raw file, then normalise structure with AI if needed
        df_raw = load_raw_file(input_path)
        df_raw = ai_normalise_structure(df_raw, input_path, client)
        total_rows = len(df_raw)

        # 2. Get a few sample rows for AI context (avoid sending entire sheet)
        sample_rows = df_raw.head(5).to_dict(orient="records")

        # 3. AI column mapping (one LLM call for the whole file)
        # Returns both the column mapping AND quality flags identifying:
        #   - columns whose values are combined address strings needing AI splitting
        #   - Odoo fields whose values contain wrong data (names in phones, cities in state, etc.)
        logger.info("Running AI column mapping …")
        mapping, flags = ai_map_columns(
            raw_columns=list(df_raw.columns),
            target_schema=schema,
            data_type=data_type,
            client=client,
            sample_rows=sample_rows,
        )
        logger.info("Column mapping: %s", json.dumps(mapping, indent=2))
        logger.info("Quality flags: %s", json.dumps(flags, indent=2))

        # 4. Rule-based Pandas cleaning (pre-rename for phone/email/name heuristics)
        df_clean = rule_based_clean(df_raw.copy(), data_type)

        # 5. Apply column mapping + AI-driven address splitting + field cleaning
        df_mapped = apply_column_mapping(df_clean, mapping, schema, data_type, client, flags=flags)

        # 6. Apply rule-based cleaning again on renamed columns for consistency
        df_mapped = rule_based_clean(df_mapped, data_type)

        # 8. Validate (rows stay in place — failures are flagged, not removed)
        df_final, df_errors, error_positions = validate_and_split(df_mapped, data_type)
        clean_count = len(df_final) - len(df_errors)

        # 9. Write output — one workbook, two sheets: "Data" (everything, with
        # failing rows highlighted red) and "Errors" (the same failing rows
        # plus the reason, for reference). Keeps every input row reachable in
        # the single file the frontend already serves, instead of silently
        # dropping rows into a second file nothing ever exposes for download.
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_final.to_excel(writer, sheet_name="Data", index=False)
            if not df_errors.empty:
                df_errors.to_excel(writer, sheet_name="Errors", index=False)

            if error_positions:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ws = writer.sheets["Data"]
                for pos in error_positions:
                    excel_row = pos + 2   # +1 for header, +1 for 1-indexing
                    for cell in ws[excel_row]:
                        cell.fill = fill

        logger.info("Output written to '%s'", output_path)

        status = "success" if df_errors.empty else "partial"
        return {
            "status": status,
            "message": f"Processed {total_rows} rows → {clean_count} clean, {len(df_errors)} errors.",
            "output_path": str(output_path),
            "stats": {
                "total": total_rows,
                "clean": clean_count,
                "errors": len(df_errors),
            },
        }

    except Exception as exc:
        logger.exception("Pipeline failed: %s", exc)
        return {
            "status": "error",
            "message": str(exc),
            "output_path": None,
            "stats": {"total": 0, "clean": 0, "errors": 0},
        }


# ===========================================================================
# 9. CLI ENTRY POINT
# ===========================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Odoo Data Migration Engine")
    parser.add_argument("input",       help="Path to raw input file (.xlsx / .csv)")
    parser.add_argument("data_type",   choices=["customer", "vendor"], help="Record type")
    parser.add_argument("--output",    default=None, help="Output file path (optional)")
    parser.add_argument("--api-key",   default=None, help="Anthropic API key (or set ANTHROPIC_API_KEY env var)")
    parser.add_argument("--batch-size", type=int, default=20, help="Address parse batch size (default: 20)")
    args = parser.parse_args()

    result = process_file(
        input_path=args.input,
        data_type=args.data_type,
        output_path=args.output,
        api_key=args.api_key,
        address_batch_size=args.batch_size,
    )

    print(json.dumps(result, indent=2))
